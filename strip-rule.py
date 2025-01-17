import json5 as json
from openpyxl import load_workbook

rule_path = 'rules/rule.json'
rule_strip_path = 'rules/rule-strip.json'
extra_cmd_path = 'extra-cmd.bat'

# load pkg defs
strip_list = [ ]
allow_list = [ ]

try:
    workbook = load_workbook('pkg-info.xlsx')
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=1, max_col=3, values_only=True):
        if row[2] == 1:
            strip_list.append(row[0])
        elif row[2] == 0:
            allow_list.append(row[0])
except Exception as e:
    print(f"Error loading Excel data: {e}")
    exit(-1)

with open(rule_path, 'r', encoding='utf-8') as file:
    rule = json.load(file)
    app_list = rule['apps']
    app_list_strip = list(filter(lambda x: x['id'] not in strip_list, app_list))
    rule['apps'] = app_list_strip

    with open(rule_strip_path, 'w', encoding='utf-8', newline='\n') as strip_file:
        json.dump(rule, strip_file, indent=4, ensure_ascii=False)

    with open(extra_cmd_path, 'w', encoding='utf-8') as extra_file:
        for app in app_list_strip:
            if app['id'] in allow_list:
                continue
            extra_file.write(f'query-pkg-info.py {app["id"]}\n')
