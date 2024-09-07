import os
import pyperclip
import re
import sys
import traceback
import json5 as json

from curl_cffi import requests
from openpyxl import load_workbook

# The XLSX to save local package info
appinfo_path = 'pkg-info.xlsx'

# The proxy server
proxy = 'socks5://127.0.0.1:8087'
# UA
USER_AGENT = '"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36'


# The package name to query
pkg_name = ''
# The software name
software_name = ''
# The software dict list
software_list = [ ]

# Dump local package info
def dump_local_app_info(workbook):
    # Dump once
    dump_dict = { }
    for root, _, files in os.walk('rules'):
        for filename in files:
            file_path = os.path.join(root, filename)
            if os.path.splitext(filename)[1] != '.json5':
                continue
            print(f'Loading {file_path} ... ', end='', flush=True)
            with open(file_path, 'r', encoding='utf-8') as file:
                rule = json.load(file)
                for item in rule['apps']:
                    pkg_name = item['id']
                    if 'name' not in item or pkg_name in software_list:
                        # No name defined or exist
                        continue
                    # Use the longer name
                    app_name = item['name']
                    if pkg_name not in dump_dict or len(app_name) > len(dump_dict[pkg_name]):
                        dump_dict[pkg_name] = app_name
                print(f'Done')
    # Save to the local file
    for key in dump_dict.keys():
        last_row = worksheet.max_row + 1
        data_to_append = [key, dump_dict[key]]
        worksheet.append(data_to_append)
    workbook.save(appinfo_path)

# Get app info from xiaomi market
def get_app_info_from_xiaomi_market(pkg_name):
    remote_url = f'https://app.mi.com/details?id={pkg_name}'
    response = requests.get(remote_url)
    app_info = re.search(r'h3 style[^>]*>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip()
    else:
        return ''

# Get app info from coolapk market
def get_app_info_from_coolapk_market(pkg_name):
    remote_url = f'https://www.coolapk.com/apk/{pkg_name}'
    headers = {
        "User-Agent": USER_AGENT,
    }
    response = requests.get(remote_url, headers=headers)
    app_info = re.search(r'detail_app_title[^>]*>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip()
    else:
        return ''

# Get app info from tencent market
def get_app_info_from_tencent_market(pkg_name):
    remote_url = f'https://sj.qq.com/appdetail/{pkg_name}'
    response = requests.get(remote_url)
    app_info = re.search(r'h1 title[^"]*"([^"]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip()
    else:
        return ''

# Get app info from apkpure market
def get_app_info_from_apkpure_market(pkg_name):
    remote_url = f'https://apkpure.com/cn/{pkg_name}'
    print(remote_url)
    headers = {
        'Host': 'apkpure.com',
        "User-Agent": USER_AGENT,
    }
    response = requests.get(remote_url, headers=headers, proxies=({'https': proxy}))
    print(response)
    app_info = re.search(r'<h1>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip()
    else:
        return ''

# Get app info from google play market
def get_app_info_from_google_play(pkg_name):
    remote_url = f'https://play.google.com/store/apps/details?id={pkg_name}'
    response = requests.get(remote_url, proxies=({'https': proxy}))
    app_info = re.search(r'main-title[^>]*>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip().split('-')[0].strip()
    else:
        return ''

# Get app info from samsung galaxy market
def get_app_info_from_galaxy_market(pkg_name):
    remote_url = f'https://galaxystore.samsung.com/detail/{pkg_name}'
    response = requests.get(remote_url, proxies=({'https': proxy}))
    app_info = re.search(r'<title>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip().split('-')[0].strip()
    else:
        return ''

# Get app info from APKsHub market
def get_app_info_from_apkshub_market(pkg_name):
    remote_url = f'https://www.apkshub.com/app/{pkg_name}'
    response = requests.get(remote_url, proxies=({'https': proxy}))
    app_info = re.search(r'App Name</span><span>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip().split('-')[0].strip()
    else:
        return ''

# Get app info from APKCombo market
def get_app_info_from_apkcombo_market(pkg_name):
    remote_url = f'https://apkcombo.com/{pkg_name}'
    response = requests.get(remote_url, headers=headers, proxies=({'https': proxy}))
    app_info = re.search(r'<p><span>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip().split('-')[0].strip()
    else:
        return ''

# Get app info from F-Droid market
def get_app_info_from_fdroid_market(pkg_name):
    remote_url = f'https://f-droid.org/packages/{pkg_name}'
    response = requests.get(remote_url, proxies=({'https': proxy}))
    app_info = re.search(r'<title>([^<]*)', response.text, flags=re.M|re.I)

    if app_info:
        return app_info.group(1).strip().split('|')[0].strip()
    else:
        return ''

# Check if the package name is not valid
def is_pkg_name_valid(software_name):
    if not software_name:
        return False
    elif '404 Page Not Found' in software_name:
        return False
    else:
        return True

# Get app info from markets
def get_app_info_from_markets(pkg_name):
    software_name = ''

    # Get app info from markets one by one
    if not is_pkg_name_valid(software_name):
        software_name = get_app_info_from_tencent_market(pkg_name)
    if not is_pkg_name_valid(software_name):
        software_name = get_app_info_from_coolapk_market(pkg_name)
    if not is_pkg_name_valid(software_name):
        software_name = get_app_info_from_xiaomi_market(pkg_name)
    # if not is_pkg_name_valid(software_name):
    #     software_name = get_app_info_from_google_play(pkg_name)
    # if not is_pkg_name_valid(software_name):
    #     software_name = get_app_info_from_galaxy_market(pkg_name)
    # if not is_pkg_name_valid(software_name):
    #     software_name = get_app_info_from_apkshub_market(pkg_name)
    # if not is_pkg_name_valid(software_name):
    #     software_name = get_app_info_from_fdroid_market(pkg_name)
    # Blocked
    # if not is_pkg_name_valid(software_name):
    #     software_name = get_app_info_from_apkcombo_market(pkg_name)
    # 403 blocked
    # if not is_pkg_name_valid(software_name):
    #     software_name = get_app_info_from_apkpure_market(pkg_name)

    # Not found finally
    if not is_pkg_name_valid(software_name):
        software_name = '?'
    else:
        pyperclip.copy(software_name)

    return software_name


if __name__ == '__main__':
    # Get package name from command line
    if len(sys.argv) > 1:
        pkg_name = sys.argv[1]
        pkg_name.strip().strip(',').strip('"')
    if not pkg_name:
        print('Usage: query-pkg-info <package_name>')
        exit()

    try:
        # First query the package name in local
        pkg_name = pkg_name.strip().strip(',').strip('"')
        workbook = load_workbook(appinfo_path)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] not in software_list:
                software_list.append(row[0])
            if row[0] == pkg_name and row[1] != '?':
                software_name = row[1]

        # If local specified
        if pkg_name == 'local':
            dump_local_app_info(workbook)
            software_name = 'Done'
        # If not found, query from the Internet
        elif not software_name:
            print('Querying...')
            software_name = get_app_info_from_markets(pkg_name)

            # Save to the local file
            last_row = worksheet.max_row + 1
            data_to_append = [pkg_name, software_name]
            worksheet.append(data_to_append)
            workbook.save(appinfo_path)

        # Print information
        print(f'Package name: {pkg_name}')
        if software_name != '?':
            print(f'Software name: {software_name}')
        else:
            print(f'Software not found')
    except:
        print('Query or save failed.')
        traceback.print_exc()
        os.system('pause')
        exit()
