import execjs
import json
import os
from openpyxl import load_workbook

extra_cmd_path = 'extra-cmd.bat'

words = [
    '红包', '权限', '兴趣', '更新', '输入法', '签到', '菜单', '二次弹窗', '屏蔽原因',
    '推荐', '自动授权'
]

# update 3rdparty rules
os.system(r'curl -o 3rdparty/AIsouler_gkd.json -L https://registry.npmmirror.com/@aisouler/gkd_subscription/latest/files/dist/AIsouler_gkd.json5')
os.system(r'curl -o 3rdparty/Adpro_gkd.json -L https://registry.npmmirror.com/@adpro/gkd_subscription/latest/files/dist/Adpro_gkd.json5')
os.system(r'del extra-cmd.bat')

# load pkg defs
strip_list = [ ]
allow_list = [ ]
workbook = load_workbook('pkg-info.xlsx')
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, max_col=3, values_only=True):
    if row[2] == 1:
        strip_list.append(row[0])
    elif row[2] == 0:
        allow_list.append(row[0])

# process one by one
for filename in os.listdir('3rdparty'):
    file_path = os.path.join('3rdparty', filename)
    if os.path.isfile(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            rule = execjs.eval(file.read())

            # remove unused info - info part
            del rule['checkUpdateUrl']
            del rule['supportUri']
            del rule['categories']

            # remove unused info - global part
            global_rule_cnt = len(rule['globalGroups'])
            for i in range(global_rule_cnt - 1, -1, -1):
                if rule['globalGroups'][i]['name'] == '更新提示':
                    del rule['globalGroups'][i]
                    continue
                rule['globalGroups'][i]['apps'] = list(filter(lambda x: x['id'] not in strip_list, rule['globalGroups'][i]['apps']))

            # remove unused info - apps part
            app_rule_cnt = len(rule['apps'])
            for i in range(app_rule_cnt - 1, -1, -1):
                if rule['apps'][i]['id'] in strip_list:
                    del rule['apps'][i]
                    continue
                rule['apps'][i]['groups'] = list(filter(lambda x: all(word not in x['name'] and (word not in x['desc'] if 'desc' in x else True) for word in words), rule['apps'][i]['groups']))

            output_path = os.path.join('rules', filename)
            with open(output_path, 'w', encoding='utf-8') as norm_file:
                json.dump(rule, norm_file, indent=4, ensure_ascii=False)

            with open(extra_cmd_path, 'a', encoding='utf-8') as extra_file:
                for app in rule['apps']:
                    if app['id'] in allow_list:
                        continue
                    extra_file.write(f'query-pkg-info.py {app["id"]}\n')
